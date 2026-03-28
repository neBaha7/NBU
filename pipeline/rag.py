"""
rag.py — Smart Direct-Context RAG (optimized for free tier)

Strategy:
1. Extract company name AND specific file reference from each question
2. Load ONLY the referenced file + profile.json (not all 100+ files)
3. Send compact context to Gemini 2.0 Flash
4. Parse structured JSON answer

This keeps context under 50k chars per question, respecting free-tier limits.
"""

import os
import re
import json
import time
from typing import List, Dict, Any, Optional, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None

from google import genai
from google.genai import types
from dotenv import load_dotenv
from tqdm import tqdm

load_dotenv()

LLM_MODEL = os.getenv('LLM_MODEL', 'gemini-2.0-flash')
QUESTIONS_FILE = os.getenv('QUESTIONS_FILE', './docs/questions_public.xlsx')
DATASET_DIR = os.getenv('DATASET_DIR', './dataset')

client = genai.Client(api_key=os.getenv('GEMINI_API_KEY'))


SYSTEM_PROMPT = """You are a financial data extraction assistant for Uzbek companies.

RULES:
1. Answer ONLY from the provided context. Do NOT use external knowledge.
2. Be EXTREMELY precise with numbers — copy exactly from source.
3. If answer_type is "int": return ONLY a plain integer (no commas/spaces/units).
4. If answer_type is "str": return the exact text string from source.
5. If not found: answer "Информация не найдена".

Respond in JSON: {"answer": <value>, "sources": [{"document_name": "file", "page_number": 1}]}
"""


def load_questions(filepath: str = QUESTIONS_FILE) -> List[Dict[str, Any]]:
    """Load questions from Excel."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    header = [str(cell.value).strip().lower() if cell.value else f'col_{i}'
              for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        q = {
            'id': int(d.get('id', 0)),
            'block': str(d.get('block', '')),
            'question': str(d.get('full_question', '')),
            'answer_type': str(d.get('answer_type', 'str')),
        }
        if q['question'] and q['question'] != 'None':
            questions.append(q)
    wb.close()
    print(f"Loaded {len(questions)} questions")
    return questions


def extract_references(question: str) -> Tuple[Optional[str], Optional[str]]:
    """Extract company name and specific file reference from question text."""
    # Extract filename like NSBU_annual_2134, NSBU_quarter_25834, MSFO_annual_810, profile.json
    file_match = re.search(r'((?:NSBU|MSFO)_(?:annual|quarter)_\d+)', question)
    file_ref = file_match.group(1) if file_match else None

    # Check for profile.json reference
    if 'profile' in question.lower():
        file_ref = file_ref or 'profile'

    # Extract company name — look for quoted names
    company_match = re.search(r'"([^"]+)"', question)
    if not company_match:
        company_match = re.search(r'«([^»]+)»', question)
    if not company_match:
        company_match = re.search(r'"([^"]+)"', question)

    company_name = company_match.group(1) if company_match else None

    return company_name, file_ref


def find_company_dir(company_name: str, company_dirs: List[str]) -> Optional[str]:
    """Match company name to directory."""
    if not company_name:
        return None

    name_lower = company_name.lower().strip()

    for cdir in company_dirs:
        dirname = os.path.basename(cdir)
        searchable = dirname.replace('_', ' ').lower()

        # Exact substring match
        if name_lower in searchable or searchable in name_lower:
            return cdir

        # Key parts match
        name_parts = name_lower.split()
        dirname_parts = searchable.split()
        if len(name_parts) >= 2:
            matches = sum(1 for p in name_parts if any(p in dp for dp in dirname_parts))
            if matches >= 2:
                return cdir

    return None


def load_targeted_context(company_dir: str, file_ref: Optional[str]) -> Tuple[str, List[Dict]]:
    """Load only the specific file referenced + profile.json."""
    from pipeline.ingest import parse_profile_json, parse_xlsx, parse_pdf

    company_id = os.path.basename(company_dir)
    all_chunks = []

    # Always load profile.json
    profile_path = os.path.join(company_dir, 'profile.json')
    if os.path.exists(profile_path):
        all_chunks.extend(parse_profile_json(profile_path, company_id))

    # If a specific file is referenced, find and load only that file
    if file_ref and file_ref != 'profile':
        # Search for the file in reports/ and other dirs
        for root, dirs, files in os.walk(company_dir):
            for fn in files:
                # Match by file_ref prefix (e.g., NSBU_annual_2134 matches NSBU_annual_2134.xlsx and .pdf)
                if fn.startswith(file_ref):
                    fpath = os.path.join(root, fn)
                    ext = os.path.splitext(fn)[1].lower()
                    if ext in ('.xlsx', '.xls'):
                        all_chunks.extend(parse_xlsx(fpath, company_id))
                    elif ext == '.pdf':
                        all_chunks.extend(parse_pdf(fpath, company_id))
    else:
        # No specific file — load all but limit to profile + first few reports
        reports_dir = os.path.join(company_dir, 'reports')
        if os.path.isdir(reports_dir):
            report_files = sorted(os.listdir(reports_dir))[:10]  # max 10 files
            for fn in report_files:
                fpath = os.path.join(reports_dir, fn)
                ext = os.path.splitext(fn)[1].lower()
                if ext in ('.xlsx', '.xls'):
                    all_chunks.extend(parse_xlsx(fpath, company_id))
                elif ext == '.pdf':
                    all_chunks.extend(parse_pdf(fpath, company_id))

    # Build context text
    parts = []
    for chunk in all_chunks:
        parts.append(chunk['text_content'])

    context = '\n\n'.join(parts)

    # Hard limit: 100k chars to stay within token limits
    if len(context) > 100000:
        context = context[:100000] + "\n[...truncated...]"

    return context, all_chunks


def query_llm(question: str, context: str, answer_type: str, attempt: int = 0) -> Dict[str, Any]:
    """Send question + targeted context to Gemini."""
    user_prompt = f"""CONTEXT:
{context}

QUESTION: {question}
ANSWER TYPE: {answer_type}

Respond in JSON format only."""

    try:
        response = client.models.generate_content(
            model=LLM_MODEL,
            contents=user_prompt,
            config=types.GenerateContentConfig(
                system_instruction=SYSTEM_PROMPT,
                temperature=0.0,
                max_output_tokens=500,
                response_mime_type="application/json",
            ),
        )

        result = json.loads(response.text)

        # Type coerce
        answer = result.get('answer', '')
        if answer_type == 'int':
            try:
                cleaned = str(answer).replace(',', '').replace(' ', '').replace('\u00a0', '')
                answer = int(float(cleaned))
            except (ValueError, TypeError):
                pass
        elif answer_type == 'float':
            try:
                cleaned = str(answer).replace(',', '').replace(' ', '').replace('\u00a0', '')
                answer = float(cleaned)
            except (ValueError, TypeError):
                pass

        result['answer'] = answer
        return result

    except Exception as e:
        if ('429' in str(e) or 'RESOURCE_EXHAUSTED' in str(e)) and attempt < 5:
            wait = 60 * (attempt + 1)
            print(f"    Rate limited (attempt {attempt+1}), waiting {wait}s...")
            time.sleep(wait)
            return query_llm(question, context, answer_type, attempt + 1)
        print(f"    [ERROR]: {str(e)[:100]}")
        return {"answer": "Error", "sources": []}


def answer_all_questions(questions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Answer all questions using targeted direct-context RAG."""
    company_dirs = sorted([
        os.path.join(DATASET_DIR, d)
        for d in os.listdir(DATASET_DIR)
        if os.path.isdir(os.path.join(DATASET_DIR, d))
    ])
    print(f"Available companies: {len(company_dirs)}")

    results = []
    for q in tqdm(questions, desc="Answering"):
        qid = q['id']
        question = q['question']
        answer_type = q['answer_type']

        # Extract company and file reference from question
        company_name, file_ref = extract_references(question)
        print(f"\n  Q{qid}: {question[:70]}...")
        print(f"    Company: {company_name} | File: {file_ref}")

        # Find company directory
        company_dir = find_company_dir(company_name, company_dirs)
        if not company_dir:
            print(f"    [WARN] Company not found!")
            results.append({'question_id': qid, 'answer': 'Информация не найдена', 'relevant_chunks': []})
            continue

        print(f"    Dir: {os.path.basename(company_dir)}")

        # Load only targeted context
        context, chunks = load_targeted_context(company_dir, file_ref)
        print(f"    Context: {len(context)} chars, {len(chunks)} chunks")

        # Query LLM
        result = query_llm(question, context, answer_type)
        answer = result.get('answer', '')
        print(f"    Answer: {str(answer)[:80]}")

        # Format sources
        relevant_chunks = []
        for src in result.get('sources', []):
            relevant_chunks.append({
                'document_name': src.get('document_name', ''),
                'page_number': src.get('page_number', 1),
            })
        if not relevant_chunks and chunks:
            relevant_chunks.append({
                'document_name': chunks[0].get('filename', ''),
                'page_number': 1,
            })

        results.append({
            'question_id': qid,
            'answer': answer,
            'relevant_chunks': relevant_chunks,
        })

        # Respect rate limits — wait between queries
        time.sleep(5)

    return results


if __name__ == '__main__':
    questions = load_questions()
    results = answer_all_questions(questions)
    with open('answers_raw.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\nSaved {len(results)} answers")
