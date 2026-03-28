"""
ingest.py — Document Ingestion Module (v2 — adapted to actual NBU dataset)

Parses all documents from dataset/:
- profile.json (company metadata — every company has this)
- reports/*.xlsx and reports/*.pdf (NSBU/MSFO financial reports — some companies)
- facts/ directory PDFs (corporate disclosures — some companies)

Each document produces chunks with metadata:
  { company_id, filename, page_or_sheet, text_content, doc_type }
"""

import os
import json
import glob
from pathlib import Path
from typing import List, Dict, Any

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from tqdm import tqdm


def parse_profile_json(filepath: str, company_id: str) -> List[Dict[str, Any]]:
    """Parse company profile.json and return it as a searchable chunk."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Build a human-readable text version of the profile
    lines = [f"=== Company Profile: {company_id} ==="]
    
    # Key fields
    field_labels = {
        'full_name_text': 'Full Name',
        'short_name_text': 'Short Name', 
        'exchange_ticket_name': 'Ticker',
        'inn': 'INN (Tax ID)',
        'location': 'Location',
        'address': 'Address',
        'email': 'Email',
        'web_site': 'Website',
        'serving_bank': 'Serving Bank',
        'account_number': 'Account Number',
        'mfo': 'MFO',
        'gov_reg_number': 'Gov Registration Number',
        'okpo': 'OKPO',
        'okonx': 'OKONX',
        'oked': 'OKED',
        'soato': 'SOATO',
        'region': 'Region',
        'status_from_stat_uz': 'Status',
        'legal_address': 'Legal Address',
        'postal_address': 'Postal Address',
    }
    
    for key, label in field_labels.items():
        val = data.get(key)
        if val is not None and str(val).strip():
            lines.append(f"{label}: {val}")
    
    # Detail info
    detail = data.get('detailinfo', {})
    if detail:
        for key in ['director_name', 'accountant_name', 'phone_number',
                     'short_info_ru', 'short_info_uz', 'short_info_en']:
            val = detail.get(key)
            if val and str(val).strip():
                lines.append(f"{key}: {val}")

    text = '\n'.join(lines)
    
    return [{
        'company_id': company_id,
        'filename': 'profile.json',
        'page_or_sheet': 'profile',
        'page_number': 1,
        'text_content': text,
        'doc_type': 'profile',
        'source_path': filepath,
    }]


def parse_xlsx(filepath: str, company_id: str) -> List[Dict[str, Any]]:
    """Parse an XLSX file. Each sheet becomes one or more chunks."""
    chunks = []
    filename = os.path.basename(filepath)

    if openpyxl is None and pd is None:
        print(f"  [WARN] No XLSX parser available, skipping {filename}")
        return chunks

    try:
        if openpyxl:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    str_row = [str(cell) if cell is not None else '' for cell in row]
                    rows.append(str_row)

                if not rows:
                    continue

                text_lines = [f"=== File: {filename} | Sheet: {sheet_name} ==="]
                for row_data in rows:
                    if all(c == '' for c in row_data):
                        continue
                    text_lines.append(' | '.join(row_data))

                text = '\n'.join(text_lines)
                if len(text.strip()) > 10:
                    chunks.append({
                        'company_id': company_id,
                        'filename': filename,
                        'page_or_sheet': sheet_name,
                        'page_number': sheet_idx,
                        'text_content': text,
                        'doc_type': 'xlsx_report',
                        'source_path': filepath,
                    })
            wb.close()
        elif pd:
            xls = pd.ExcelFile(filepath)
            for sheet_idx, sheet_name in enumerate(xls.sheet_names, start=1):
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                lines = [f"=== File: {filename} | Sheet: {sheet_name} ==="]
                for _, row in df.iterrows():
                    vals = [str(v) if pd.notna(v) else '' for v in row]
                    if any(v != '' for v in vals):
                        lines.append(' | '.join(vals))
                text = '\n'.join(lines)
                if len(text.strip()) > 10:
                    chunks.append({
                        'company_id': company_id,
                        'filename': filename,
                        'page_or_sheet': sheet_name,
                        'page_number': sheet_idx,
                        'text_content': text,
                        'doc_type': 'xlsx_report',
                        'source_path': filepath,
                    })
    except Exception as e:
        print(f"  [ERROR] Failed to parse {filename}: {e}")

    return chunks


def parse_pdf(filepath: str, company_id: str, doc_type: str = 'pdf_report') -> List[Dict[str, Any]]:
    """Parse a PDF file page by page using pdfplumber."""
    chunks = []
    filename = os.path.basename(filepath)

    if pdfplumber is None:
        print(f"  [WARN] pdfplumber not available, skipping {filename}")
        return chunks

    try:
        with pdfplumber.open(filepath) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ''

                # Also extract tables
                tables = page.extract_tables() or []
                for table in tables:
                    if table:
                        table_lines = []
                        for row in table:
                            str_row = [str(cell) if cell else '' for cell in row]
                            table_lines.append(' | '.join(str_row))
                        text += '\n[TABLE]\n' + '\n'.join(table_lines) + '\n[/TABLE]\n'

                if text.strip():
                    chunks.append({
                        'company_id': company_id,
                        'filename': filename,
                        'page_or_sheet': f'page_{page_num}',
                        'page_number': page_num,
                        'text_content': f"=== File: {filename} | Page {page_num} ===\n{text}",
                        'doc_type': doc_type,
                        'source_path': filepath,
                    })
    except Exception as e:
        print(f"  [ERROR] pdfplumber failed for {filename}: {e}")

    return chunks


def ingest_company(company_dir: str) -> List[Dict[str, Any]]:
    """Ingest all documents for a single company."""
    company_id = os.path.basename(company_dir)
    all_chunks = []

    # 1. Parse profile.json
    profile_path = os.path.join(company_dir, 'profile.json')
    if os.path.exists(profile_path):
        all_chunks.extend(parse_profile_json(profile_path, company_id))

    # 2. Parse reports/ (XLSX + PDF)
    reports_dir = os.path.join(company_dir, 'reports')
    if os.path.isdir(reports_dir):
        for fpath in sorted(glob.glob(os.path.join(reports_dir, '*'))):
            if not os.path.isfile(fpath):
                continue
            ext = os.path.splitext(fpath)[1].lower()
            if ext in ('.xlsx', '.xls'):
                all_chunks.extend(parse_xlsx(fpath, company_id))
            elif ext == '.pdf':
                all_chunks.extend(parse_pdf(fpath, company_id, 'pdf_report'))

    # 3. Parse facts/ and any other subdirs with PDFs
    for subdir_name in os.listdir(company_dir):
        subdir_path = os.path.join(company_dir, subdir_name)
        if os.path.isdir(subdir_path) and subdir_name != 'reports':
            for root, _, fnames in os.walk(subdir_path):
                for fn in sorted(fnames):
                    fpath = os.path.join(root, fn)
                    ext = os.path.splitext(fn)[1].lower()
                    if ext == '.pdf':
                        all_chunks.extend(parse_pdf(fpath, company_id, 'fact_disclosure'))
                    elif ext in ('.xlsx', '.xls'):
                        all_chunks.extend(parse_xlsx(fpath, company_id))

    return all_chunks


def ingest_all(dataset_dir: str) -> List[Dict[str, Any]]:
    """Ingest all companies from the dataset directory."""
    all_chunks = []
    
    # Get company directories (skip json files)
    company_dirs = sorted([
        os.path.join(dataset_dir, d) 
        for d in os.listdir(dataset_dir)
        if os.path.isdir(os.path.join(dataset_dir, d))
    ])

    print(f"Found {len(company_dirs)} companies in {dataset_dir}")

    for cdir in tqdm(company_dirs, desc="Ingesting companies"):
        cid = os.path.basename(cdir)
        chunks = ingest_company(cdir)
        if chunks:
            print(f"  {cid}: {len(chunks)} chunks")
        all_chunks.extend(chunks)

    print(f"\nTotal chunks ingested: {len(all_chunks)}")
    return all_chunks


if __name__ == '__main__':
    import sys
    dataset_dir = sys.argv[1] if len(sys.argv) > 1 else './dataset'
    chunks = ingest_all(dataset_dir)
    with open('raw_chunks.json', 'w', encoding='utf-8') as f:
        json.dump(chunks, f, ensure_ascii=False, indent=2)
    print(f"Saved {len(chunks)} chunks to raw_chunks.json")
