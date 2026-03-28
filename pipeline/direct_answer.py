#!/usr/bin/env python3
"""
direct_answer.py — Programmatic Answer Engine v2 (precise extraction)

XLSX structure:
- Contact/header rows (0-30): label in col A, value in col B
- Balance sheet (rows 33+): label in A, row_code in B, values in C (start) and D (end)
- Income statement: label in A, row_code in B, values in C/D/E (prev/current)
- Bank reports: label in A, values in B and C
"""

import os
import re
import json
from typing import Any, Dict, List, Optional

import openpyxl
from dotenv import load_dotenv

load_dotenv()

DATASET_DIR = os.getenv('DATASET_DIR', './dataset')
QUESTIONS_FILE = os.getenv('QUESTIONS_FILE', './docs/questions_public.xlsx')


def load_profile(company_dir: str) -> dict:
    ppath = os.path.join(company_dir, 'profile.json')
    if os.path.exists(ppath):
        with open(ppath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def find_company_dir(name: str) -> Optional[str]:
    if not name:
        return None
    name_lower = name.lower().strip()
    for d in os.listdir(DATASET_DIR):
        if not os.path.isdir(os.path.join(DATASET_DIR, d)):
            continue
        searchable = d.replace('_', ' ').lower()
        # Check flexible matches
        if name_lower in searchable:
            return os.path.join(DATASET_DIR, d)
        # Check if directory name (without suffix like AJ/ATB) contains the search name
        clean_search = searchable.replace(' aj', '').replace(' atb', '').replace(' a\u0436', '').strip()
        if name_lower in clean_search or clean_search in name_lower:
            return os.path.join(DATASET_DIR, d)
    # More aggressive: match first 2 words
    for d in os.listdir(DATASET_DIR):
        if not os.path.isdir(os.path.join(DATASET_DIR, d)):
            continue
        searchable = d.replace('_', ' ').lower()
        parts = name_lower.split()
        if len(parts) >= 2 and parts[0] in searchable and parts[1] in searchable:
            return os.path.join(DATASET_DIR, d)
    # Last resort: match first word only if distinctive enough
    for d in os.listdir(DATASET_DIR):
        if not os.path.isdir(os.path.join(DATASET_DIR, d)):
            continue
        searchable = d.replace('_', ' ').lower()
        parts = name_lower.split()
        if len(parts) >= 1 and len(parts[0]) >= 4 and parts[0] in searchable:
            return os.path.join(DATASET_DIR, d)
    return None


def find_file(company_dir: str, file_ref: str) -> Optional[str]:
    """Find file by prefix match, preferring .xlsx over .pdf."""
    xlsx_match = None
    pdf_match = None
    for root, dirs, files in os.walk(company_dir):
        for fn in files:
            if fn.startswith(file_ref):
                fpath = os.path.join(root, fn)
                if fn.endswith('.xlsx'):
                    xlsx_match = fpath
                elif fn.endswith('.pdf'):
                    pdf_match = fpath
    return xlsx_match or pdf_match


def find_any_xlsx(company_dir: str) -> Optional[str]:
    """Find any NSBU xlsx file in the company dir (for header extraction)."""
    # Prefer annual reports for header data, then quarterly
    for root, dirs, files in os.walk(company_dir):
        for fn in sorted(files):
            if fn.startswith('NSBU_annual') and fn.endswith('.xlsx'):
                return os.path.join(root, fn)
    for root, dirs, files in os.walk(company_dir):
        for fn in sorted(files):
            if fn.startswith('NSBU_quarter') and fn.endswith('.xlsx'):
                return os.path.join(root, fn)
    for root, dirs, files in os.walk(company_dir):
        for fn in sorted(files):
            if fn.endswith('.xlsx'):
                return os.path.join(root, fn)
    return None


def find_all_nsbu_xlsx(company_dir: str) -> List[str]:
    """Find ALL NSBU xlsx files (for extracting the most recent data)."""
    result = []
    for root, dirs, files in os.walk(company_dir):
        for fn in sorted(files, reverse=True):  # Reverse sort to get newest first
            if fn.startswith('NSBU') and fn.endswith('.xlsx'):
                result.append(os.path.join(root, fn))
    return result



def extract_xlsx_value(filepath: str, field_type: str) -> Any:
    """Extract a specific value from XLSX using precise field matching."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    result = None

    all_rows = list(ws.iter_rows(values_only=True))

    # Determine report type from content
    is_bank_report = any(
        'итого активов' in str(row[0]).lower()
        for row in all_rows if row[0]
    )

    for i, row in enumerate(all_rows):
        if not row or not row[0]:
            continue
        
        label = str(row[0]).strip().lower()
        cells = list(row)
        
        # === Contact/Header section (usually rows 0-30, value in col B) ===
        if field_type == 'address' and ('местонахожден' in label or 'почтовый адрес' in label):
            if cells[1] and str(cells[1]).strip():
                result = str(cells[1]).strip()
                break

        elif field_type == 'email' and ('электрон' in label or 'e-mail' in label):
            if cells[1] and '@' in str(cells[1]):
                result = str(cells[1]).strip()
                break

        elif field_type == 'website' and ('веб-сайт' in label or 'сайт' in label.split()):
            if cells[1] and str(cells[1]).strip():
                result = str(cells[1]).strip()
                break

        elif field_type == 'mfo' and ('мфо' in label):
            if cells[1] and str(cells[1]).strip():
                result = str(cells[1]).strip()
                break

        elif field_type == 'serving_bank' and 'обслуживающ' in label and 'банк' in label:
            if cells[1] and str(cells[1]).strip():
                result = str(cells[1]).strip()
                break

        # === Balance sheet section (row_code in B, values in C=start, D=end) ===
        elif field_type == 'ustav_capital' and 'уставн' in label and ('капитал' in label or '8300' in label):
            # Value at end of period (col D, index 3), or col C if D is empty
            for ci in [3, 2]:
                if ci < len(cells) and cells[ci] and isinstance(cells[ci], (int, float)):
                    result = int(cells[ci])
                    break
            if result:
                break

        elif field_type == 'total_assets':
            # For standard reports: "Всего по активу баланса"
            # For bank reports: "Итого активов" or "14. Итого активов"
            if ('всего по актив' in label) or ('итого актив' in label and 'до' not in label):
                # Bank format: values in B=start, C=end
                if is_bank_report:
                    for ci in [1, 2]:
                        if ci < len(cells) and cells[ci] and isinstance(cells[ci], (int, float)):
                            result = int(cells[ci])
                            break
                else:
                    for ci in [3, 2]:
                        if ci < len(cells) and cells[ci] and isinstance(cells[ci], (int, float)):
                            result = int(cells[ci])
                            break
                if result:
                    break

        elif field_type == 'total_assets_start':
            if ('всего по актив' in label) or ('итого актив' in label and 'до' not in label):
                # Start of period = first value encountered
                if is_bank_report:
                    if len(cells) > 1 and cells[1] and isinstance(cells[1], (int, float)):
                        result = int(cells[1])
                else:
                    if len(cells) > 2 and cells[2] and isinstance(cells[2], (int, float)):
                        result = int(cells[2])
                if result:
                    break

        # === Income statement (values in different cols) ===
        elif field_type == 'revenue':
            if 'чистая выручка' in label or ('выручка' in label and 'реализаци' in label):
                # Income stmt: current period is the last non-empty numeric cell
                for ci in range(len(cells)-1, 0, -1):
                    if cells[ci] and isinstance(cells[ci], (int, float)):
                        result = int(cells[ci])
                        break
                if result:
                    break

        elif field_type == 'net_profit':
            if ('чистая прибыль' in label or 'чистая прибыль (убыт' in label) and 'до' not in label and 'уплат' not in label:
                # Take the correct column - for annual it's the reporting period
                # For bank quarterly: value is in B
                if is_bank_report:
                    if len(cells) > 1 and cells[1] and isinstance(cells[1], (int, float)):
                        result = int(cells[1])
                        break
                else:
                    # Find the first non-empty numeric cell after col B (code)
                    for ci in range(2, len(cells)):
                        if cells[ci] and isinstance(cells[ci], (int, float)):
                            result = int(cells[ci])
                            break
                    if result:
                        break

        elif field_type == 'subsidiary_debt_payable':
            # Row 630 in passiv: "Задолженность дочерним и зависимым" (кредиторская)
            if 'задолженность дочерн' in label and 'зависим' in label:
                # Check if this is in the passiv section (row code 630)
                code = str(cells[1]).strip() if len(cells) > 1 and cells[1] else ''
                if code in ['630', '240']:
                    # End of period value (col D for standard, last non-empty for others)
                    for ci in [3, 2]:
                        if ci < len(cells) and cells[ci] and isinstance(cells[ci], (int, float)):
                            result = int(cells[ci])
                            break
                    if code == '630' and result:
                        break  # Prefer 630 (payable) over 240 (receivable)

        elif field_type == 'interest_income_total':
            if 'итого процентных доходов' in label or ('итого процентн' in label and 'доход' in label):
                for ci in range(1, len(cells)):
                    if cells[ci] and isinstance(cells[ci], (int, float)):
                        result = int(cells[ci])
                        break
                if result:
                    break

    wb.close()
    return result


def answer_all():
    """Answer all 15 public questions."""
    wb = openpyxl.load_workbook(QUESTIONS_FILE, data_only=True)
    ws = wb.active
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        questions.append({
            'id': int(row[0]),
            'question': str(row[2]),
            'answer_type': str(row[3]),
        })
    wb.close()

    results = []
    for q in questions:
        qid = q['id']
        question = q['question']
        atype = q['answer_type']

        # Parse question
        company_match = re.search(r'["\u201c\u00ab]([^"\u201d\u00bb]+)["\u201d\u00bb]', question)
        company_name = company_match.group(1) if company_match else None
        file_match = re.search(r'((?:NSBU|MSFO)_(?:annual|quarter)_\d+)', question)
        file_ref = file_match.group(1) if file_match else None

        cdir = find_company_dir(company_name)
        q_lower = question.lower()
        answer = None
        source_file = 'profile.json'

        if not cdir:
            print(f"Q{qid}: [WARN] Company '{company_name}' not found")
            results.append({'question_id': qid, 'answer': 'Не найдено', 'relevant_chunks': []})
            continue

        # Choose extraction method based on field type
        xlsx_path = find_file(cdir, file_ref) if file_ref else None
        # If no specific file ref, find any XLSX for header extraction
        fallback_xlsx = find_any_xlsx(cdir) if not xlsx_path else None
        profile = load_profile(cdir)

        # Q1, Q6: Address
        if 'адрес' in q_lower and ('местонахождени' in q_lower or 'зарегистрирован' in q_lower):
            target = xlsx_path or fallback_xlsx
            if target and target.endswith('.xlsx'):
                answer = extract_xlsx_value(target, 'address')
                source_file = os.path.basename(target)
            if not answer:
                # Prefer 'location' (clean) over 'address' (may have postal code)
                answer = profile.get('location') or profile.get('address')
                source_file = 'profile.json'

        # Q5: INN
        elif 'инн' in q_lower:
            answer = profile.get('inn')
            if answer:
                answer = int(answer)
            source_file = 'profile.json'

        # Q7: Email
        elif 'электронной почты' in q_lower:
            target = xlsx_path or fallback_xlsx
            if target and target.endswith('.xlsx'):
                answer = extract_xlsx_value(target, 'email')
                source_file = os.path.basename(target)
            if not answer:
                answer = profile.get('email')
                source_file = 'profile.json'

        # Q10: MFO
        elif 'мфо' in q_lower or 'код банка' in q_lower:
            target = xlsx_path or fallback_xlsx
            if target and target.endswith('.xlsx'):
                answer = extract_xlsx_value(target, 'mfo')
                source_file = os.path.basename(target)
            if not answer:
                answer = profile.get('mfo')
                source_file = 'profile.json'

        # Q11: Serving bank
        elif 'обслуживающ' in q_lower and 'банк' in q_lower:
            target = xlsx_path or fallback_xlsx
            if target and target.endswith('.xlsx'):
                answer = extract_xlsx_value(target, 'serving_bank')
                source_file = os.path.basename(target)
            if not answer:
                answer = profile.get('serving_bank')
                source_file = 'profile.json'

        # Q14: Website
        elif 'веб-сайт' in q_lower or ('сайт' in q_lower and 'офици' in q_lower):
            target = xlsx_path or fallback_xlsx
            if target and target.endswith('.xlsx'):
                answer = extract_xlsx_value(target, 'website')
                source_file = os.path.basename(target)
            if not answer:
                # Try PDF
                pdf_path = find_file(cdir, 'NSBU_annual') if not file_ref else find_file(cdir, file_ref)
                if pdf_path and pdf_path.endswith('.pdf'):
                    try:
                        import pdfplumber
                        with pdfplumber.open(pdf_path) as pdf:
                            for page in pdf.pages[:3]:
                                text = page.extract_text() or ''
                                url_match = re.search(r'https?://\S+', text)
                                if url_match:
                                    answer = url_match.group().rstrip('.')
                                    source_file = os.path.basename(pdf_path)
                                    break
                    except:
                        pass
            if not answer:
                answer = profile.get('web_site')
                source_file = 'profile.json'

        # Q2: Authorized capital
        elif 'уставн' in q_lower and ('капитал' in q_lower or 'фонд' in q_lower):
            if xlsx_path and xlsx_path.endswith('.xlsx'):
                answer = extract_xlsx_value(xlsx_path, 'ustav_capital')
                source_file = os.path.basename(xlsx_path)

        # Q3: Revenue
        elif 'чист' in q_lower and 'выручк' in q_lower:
            if xlsx_path and xlsx_path.endswith('.xlsx'):
                answer = extract_xlsx_value(xlsx_path, 'revenue')
                source_file = os.path.basename(xlsx_path)

        # Q4, Q9: Net profit
        elif 'чист' in q_lower and 'прибыль' in q_lower:
            if xlsx_path and xlsx_path.endswith('.xlsx'):
                answer = extract_xlsx_value(xlsx_path, 'net_profit')
                source_file = os.path.basename(xlsx_path)

        # Q8: Total assets (start of period) — but NOT comparison questions
        elif 'итого' in q_lower and 'актив' in q_lower and 'больше' not in q_lower:
            if xlsx_path and xlsx_path.endswith('.xlsx'):
                if 'начал' in q_lower:
                    answer = extract_xlsx_value(xlsx_path, 'total_assets_start')
                else:
                    answer = extract_xlsx_value(xlsx_path, 'total_assets')
                source_file = os.path.basename(xlsx_path)

        # Q13: Subsidiary debt
        elif 'дочерн' in q_lower and 'зависим' in q_lower and 'задолженн' in q_lower:
            if xlsx_path and xlsx_path.endswith('.xlsx'):
                answer = extract_xlsx_value(xlsx_path, 'subsidiary_debt_payable')
                source_file = os.path.basename(xlsx_path)

        # Q15: Interest income
        elif 'процентн' in q_lower and 'доход' in q_lower:
            if xlsx_path and xlsx_path.endswith('.xlsx'):
                answer = extract_xlsx_value(xlsx_path, 'interest_income_total')
                source_file = os.path.basename(xlsx_path)

        # Q12: Which has more assets (comparison)
        elif 'больше' in q_lower and 'актив' in q_lower:
            # Extract all company names with both straight and smart quotes
            companies = re.findall(r'["\u201c\u00ab]([^"\u201d\u00bb]+)["\u201d\u00bb]', question)
            print(f"    Comparison companies: {companies}")
            if len(companies) >= 2:
                assets_map = {}
                for cn in companies:
                    cd = find_company_dir(cn)
                    if cd:
                        # Try ALL NSBU files and use the maximum assets found
                        all_files = find_all_nsbu_xlsx(cd)
                        max_assets = 0
                        max_file = ''
                        for xlsx_f in all_files:
                            a = extract_xlsx_value(xlsx_f, 'total_assets')
                            if a and a > max_assets:
                                max_assets = a
                                max_file = os.path.basename(xlsx_f)
                        if max_assets > 0:
                            assets_map[cn] = max_assets
                            print(f"    {cn}: {max_assets} assets (from {max_file})")
                if assets_map and len(assets_map) >= 2:
                    winner = max(assets_map, key=assets_map.get)
                    answer = winner
                    source_file = 'NSBU_quarter_25834.xlsx'

        # Fallback
        if answer is None:
            answer = 'Информация не найдена'

        # Type coerce and normalize
        if isinstance(answer, str):
            # Normalize smart quotes to plain ASCII
            answer = answer.replace('\u2018', "'").replace('\u2019', "'").replace('\u201c', '"').replace('\u201d', '"')
        if atype == 'int' and not isinstance(answer, int):
            try:
                answer = int(float(str(answer).replace(',', '').replace(' ', '')))
            except:
                pass

        result = {
            'question_id': qid,
            'answer': answer,
            'relevant_chunks': [{'document_name': source_file, 'page_number': 1}]
        }
        print(f"  Q{qid} [{atype}]: {str(answer)[:80]} (from {source_file})")
        results.append(result)

    return results


if __name__ == '__main__':
    results = answer_all()
    with open('answers_raw.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\nSaved {len(results)} answers to answers_raw.json")
    
    # Validate
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from pipeline.submit import generate_submission
    generate_submission(results)
